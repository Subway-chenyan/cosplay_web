#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ChinaJoy Cosplay视频数据解析脚本

从bilibili爬取的CSV文件中提取ChinaJoy相关视频，
解析标题中的社团名称、剧名、比赛名和年份等信息，
转换为后端可导入的Excel格式。

使用方法:
    python parse_chinajoy_videos.py
"""

import pandas as pd
import re
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ChinaJoyVideoParser:
    def __init__(self, csv_file_path, output_file_path):
        self.csv_file_path = csv_file_path
        self.output_file_path = output_file_path
        self.chinajoy_videos = []
        
    def load_csv_data(self):
        """加载CSV数据"""
        try:
            self.df = pd.read_csv(self.csv_file_path)
            print(f"成功加载CSV文件，共{len(self.df)}条记录")
            return True
        except Exception as e:
            print(f"加载CSV文件失败: {e}")
            return False
    
    def filter_chinajoy_videos(self):
        """筛选ChinaJoy相关视频"""
        chinajoy_pattern = r'ChinaJoy.*?Cosplay.*?联赛|Cosplay.*?ChinaJoy.*?联赛'
        
        # 筛选包含ChinaJoy Cosplay联赛的视频
        mask = self.df['title'].str.contains(chinajoy_pattern, case=False, na=False)
        self.chinajoy_videos = self.df[mask].copy()
        
        print(f"筛选出{len(self.chinajoy_videos)}条ChinaJoy相关视频")
        return len(self.chinajoy_videos) > 0
    
    def parse_title_info(self, title):
        """解析标题信息
        
        标题格式示例:
        【上海赛区】八十一禁动漫社--永远的任天堂2 玩家之心（2025ChinaJoy Cosplay超级联赛总决赛）
        【华东中赛区小团体】Stray Dogs--Cyberpunk：无为之城（2025ChinaJoy Cosplay超级联赛总决赛）
        
        返回:
            dict: 包含解析出的信息
        """
        result = {
            'region': '',
            'group_name': '',
            'performance_name': '',
            'competition_name': '',
            'year': '',
            'is_small_group': False
        }
        
        try:
            # 提取赛区信息
            region_match = re.search(r'【([^】]+)】', title)
            if region_match:
                region_info = region_match.group(1)
                result['region'] = region_info
                result['is_small_group'] = '小团体' in region_info
            
            # 提取社团名称和剧名
            # 匹配格式：】社团名--剧名（ 或 】【团体】社团名--剧名（
            group_performance_match = re.search(r'】(?:【团体】)?([^--]+)--([^（]+)（', title)
            if group_performance_match:
                result['group_name'] = group_performance_match.group(1).strip()
                result['performance_name'] = group_performance_match.group(2).strip()
            else:
                # 如果上面没有匹配到，尝试匹配没有括号的格式：】【团体】社团名--剧名
                group_performance_match2 = re.search(r'】(?:【团体】)?([^--]+)--(.+?)(?:\s|$)', title)
                if group_performance_match2:
                    result['group_name'] = group_performance_match2.group(1).strip()
                    result['performance_name'] = group_performance_match2.group(2).strip()
            
            # 提取年份和比赛名称
            # 先尝试匹配括号内的格式：（2025ChinaJoy Cosplay超级联赛总决赛）
            competition_match = re.search(r'（(\d{4})(.*?ChinaJoy.*?Cosplay.*?联赛[^）]*)）', title)
            if competition_match:
                result['year'] = competition_match.group(1)
                competition_full = competition_match.group(2)
                # 提取核心比赛名称
                if 'ChinaJoy Cosplay超级联赛' in competition_full:
                    result['competition_name'] = 'ChinaJoy Cosplay超级联赛'
                elif 'ChinaJoy Cosplay联赛' in competition_full:
                    result['competition_name'] = 'ChinaJoy Cosplay联赛'
                else:
                    result['competition_name'] = competition_full.strip()
            else:
                # 如果括号内没有匹配到，尝试匹配开头的年份格式：2025 ChinaJoy 或 2024ChinaJoy 或 2023万达广场XChinaJoy 或 2021.5.16ChinaJoy
                year_match = re.search(r'^(\d{4})(?:\s*|万达广场[X×]|\s+万达广场[X×]|\.\d+\.\d+)ChinaJoy', title)
                if year_match:
                    result['year'] = year_match.group(1)
                    # 提取比赛名称
                    if 'ChinaJoy Cosplay超级联赛' in title:
                        result['competition_name'] = 'ChinaJoy Cosplay超级联赛'
                    elif 'ChinaJoy Cosplay联赛' in title:
                        result['competition_name'] = 'ChinaJoy Cosplay联赛'
                    else:
                        # 尝试提取更完整的比赛名称
                        comp_match = re.search(r'ChinaJoy[^【]*?联赛[^【]*?(?=【|$)', title)
                        if comp_match:
                            result['competition_name'] = comp_match.group().strip()
                        else:
                            result['competition_name'] = 'ChinaJoy Cosplay超级联赛'  # 默认值
            
        except Exception as e:
            print(f"解析标题失败: {title}, 错误: {e}")
        
        return result
    
    def extract_region_from_title(self, region_info):
        """从赛区信息中提取省份和城市"""
        province = ''
        city = ''
        
        # 特殊地区映射
        region_mapping = {
            '上海': ('上海市', '上海市'),
            '北京': ('北京市', '北京市'),
            '重庆': ('重庆市', '重庆市'),
            '华东北': ('', ''),  # 华东北赛区覆盖多个省份
            '华东中': ('', ''),
            '华东南': ('', ''),
            '华南': ('', ''),
            '华中': ('', ''),
            '华北': ('', ''),
            '东北': ('', ''),
            '西北': ('', ''),
            '西南': ('', ''),
        }
        
        for key, (prov, ct) in region_mapping.items():
            if key in region_info:
                province = prov
                city = ct
                break
        
        return province, city
    
    def generate_tags(self, performance_name, region):
        """根据剧名和赛区生成标签"""
        tags = []
        
        # 根据剧名生成IP标签
        ip_keywords = {
            '初音未来': 'VOCALOID:IP',
            '东方': '东方Project:IP',
            '原神': '原神:IP',
            '崩坏': '崩坏系列:IP',
            '明日方舟': '明日方舟:IP',
            '剑网3': '剑网3:IP',
            '剑三': '剑网3:IP',
            '魔兽': '魔兽世界:IP',
            'Fate': 'Fate系列:IP',
            '海贼王': '海贼王:IP',
            '火影': '火影忍者:IP',
            '死神': '死神:IP',
            'Bleach': '死神:IP',
            '龙族': '龙族:IP',
            '犬夜叉': '犬夜叉:IP',
            '赛博朋克': '赛博朋克2077:IP',
            '星露谷': '星露谷物语:IP',
            '任天堂': '任天堂游戏:IP',
            '黑执事': '黑执事:IP',
            '美少女战士': '美少女战士:IP',
            '魔卡少女樱': '魔卡少女樱:IP',
            '白蛇': '白蛇传:IP',
            '不良人': '画江湖:IP',
            '成龙历险记': '成龙历险记:IP',
        }
        
        for keyword, tag in ip_keywords.items():
            if keyword in performance_name:
                tags.append(tag)
                break
        
        # 根据剧名生成风格标签
        style_keywords = {
            '古风': '古风:风格',
            '现代': '现代:风格',
            '科幻': '科幻:风格',
            '魔幻': '魔幻:风格',
            '甜美': '甜美:风格',
            '帅气': '帅气:风格',
            '可爱': '可爱:风格',
            '优雅': '优雅:风格',
        }
        
        for keyword, tag in style_keywords.items():
            if keyword in performance_name:
                tags.append(tag)
        
        # 添加比赛相关标签
        tags.append('ChinaJoy:其他')
        if '小团体' in region:
            tags.append('小团体:其他')
        else:
            tags.append('大团体:其他')
        
        return ','.join(tags)
    
    def convert_to_import_format(self):
        """转换为导入格式"""
        import_data = []
        
        for _, row in self.chinajoy_videos.iterrows():
            # 解析标题信息
            title_info = self.parse_title_info(row['title'])
            
            # 提取省份城市
            province, city = self.extract_region_from_title(title_info['region'])
            
            # 生成标签
            tags = self.generate_tags(title_info['performance_name'], title_info['region'])
            
            # 构建导入数据
            import_row = {
                'bv_number': row['bvid'],
                'title': row['title'],
                'url': row['video_url'],
                'description': row['description'] if pd.notna(row['description']) else '',
                'thumbnail': row['cover_url'] if pd.notna(row['cover_url']) else '',
                'year': title_info['year'],
                'group_name': title_info['group_name'],
                'competition_name': title_info['competition_name'],
                'tags': tags,
                'group_description': f"{title_info['group_name']}是一个专业的Cosplay社团" if title_info['group_name'] else '',
                'group_founded_date': '',  # 无法从标题获取
                'group_province': province,
                'group_city': city,
                'group_location': title_info['region'],
                'group_website': '',
                'group_email': '',
                'group_phone': '',
                'group_weibo': '',
                'group_wechat': '',
                'group_qq_group': '',
                'group_bilibili': '',
                'competition_description': f"{title_info['competition_name']}是国内知名的Cosplay比赛" if title_info['competition_name'] else '',
                'competition_website': '',
                'award_names': '',  # 无法从标题获取
                'award_years': '',
                'award_descriptions': ''
            }
            
            import_data.append(import_row)
        
        return import_data
    
    def save_to_excel(self, import_data):
        """保存为Excel文件"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "ChinaJoy视频数据"
            
            # 定义列标题
            headers = [
                'bv_number', 'title', 'url', 'description', 'thumbnail', 'year',
                'group_name', 'competition_name', 'tags', 'group_description',
                'group_founded_date', 'group_province', 'group_city', 'group_location',
                'group_website', 'group_email', 'group_phone', 'group_weibo',
                'group_wechat', 'group_qq_group', 'group_bilibili',
                'competition_description', 'competition_website', 'award_names',
                'award_years', 'award_descriptions'
            ]
            
            # 写入标题行
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 写入数据
            for row_idx, data in enumerate(import_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=data.get(header, ''))
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(self.output_file_path)
            print(f"成功保存Excel文件: {self.output_file_path}")
            print(f"共导出{len(import_data)}条记录")
            return True
            
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
            return False
    
    def generate_summary_report(self, import_data):
        """生成汇总报告"""
        print("\n=== ChinaJoy视频数据解析汇总报告 ===")
        print(f"总视频数量: {len(import_data)}")
        
        # 按年份统计
        year_stats = {}
        for data in import_data:
            year = data['year']
            year_stats[year] = year_stats.get(year, 0) + 1
        
        print("\n按年份统计:")
        for year in sorted(year_stats.keys()):
            print(f"  {year}年: {year_stats[year]}个视频")
        
        # 按社团统计
        group_stats = {}
        for data in import_data:
            group = data['group_name']
            if group:
                group_stats[group] = group_stats.get(group, 0) + 1
        
        print(f"\n社团数量: {len(group_stats)}个")
        print("社团视频数量排行（前10）:")
        sorted_groups = sorted(group_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        for group, count in sorted_groups:
            print(f"  {group}: {count}个视频")
        
        # 按比赛统计
        competition_stats = {}
        for data in import_data:
            comp = data['competition_name']
            if comp:
                competition_stats[comp] = competition_stats.get(comp, 0) + 1
        
        print(f"\n比赛类型统计:")
        for comp, count in competition_stats.items():
            print(f"  {comp}: {count}个视频")
    
    def run(self):
        """运行解析流程"""
        print("开始解析ChinaJoy Cosplay视频数据...")
        
        # 加载数据
        if not self.load_csv_data():
            return False
        
        # 筛选ChinaJoy视频
        if not self.filter_chinajoy_videos():
            print("未找到ChinaJoy相关视频")
            return False
        
        # 转换为导入格式
        import_data = self.convert_to_import_format()
        
        # 保存Excel文件
        if not self.save_to_excel(import_data):
            return False
        
        # 生成汇总报告
        self.generate_summary_report(import_data)
        
        print("\n解析完成！")
        return True


def main():
    """主函数"""
    # 文件路径配置
    csv_file = "/home/ubuntu/cosplay_web/bilibili_crawler/user_27660646_videos/user_27660646_videos.csv"
    output_file = "/home/ubuntu/cosplay_web/backend/upload_data/chinajoy_videos_import.xlsx"
    
    # 检查输入文件是否存在
    if not os.path.exists(csv_file):
        print(f"错误: CSV文件不存在 - {csv_file}")
        return
    
    # 创建解析器并运行
    parser = ChinaJoyVideoParser(csv_file, output_file)
    parser.run()


if __name__ == "__main__":
    main()